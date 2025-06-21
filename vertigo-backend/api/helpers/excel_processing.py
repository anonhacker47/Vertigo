import os
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Global Fills
green_fill = PatternFill(start_color="58d68d", end_color="58d68d", fill_type="solid")
red_fill = PatternFill(start_color="ec7063", end_color="ec7063", fill_type="solid")

def get_fill(value):
    try:
        read, total = map(int, str(value).split("/"))
        if read == 0:
            return red_fill
        elif read == total:
            return green_fill
    except:
        return None

def apply_fill_conditionally(sheet, df, columns):
    for col in columns:
        col_idx = df.columns.get_loc(col) + 1
        for row in range(2, len(df) + 2):
            cell = sheet.cell(row=row, column=col_idx)
            fill = get_fill(cell.value)
            if fill:
                cell.fill = fill

def add_thumbnails(sheet, df, cover_path):
    thumb_idx = df.columns.get_loc("Thumbnail") + 1
    thumb_letter = get_column_letter(thumb_idx)
    for row_idx, thumbnail in enumerate(df["Thumbnail"], start=2):
        cell_location = f"{thumb_letter}{row_idx}"
        sheet[cell_location].value = None
        image_path = os.path.join(cover_path, thumbnail) if thumbnail else None
        if image_path and os.path.exists(image_path):
            img = OpenpyxlImage(image_path)
            img.width = 100
            img.height = 150
            sheet.add_image(img, cell_location)
        else:
            sheet[cell_location] = "Image not found"

def parse_owned_count(value):
    try:
        if isinstance(value, str) and '/' in value:
            return int(value.split('/')[0].strip())
        elif isinstance(value, (int, float)):  # fallback if value is already a number
            return int(value)
    except:
        pass
    return 0  # default fallback


def set_sheet_dimensions(sheet, df, row_height=120, col_width=20):
    for i, _ in enumerate(df.columns, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = col_width
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        sheet.row_dimensions[row[0].row].height = row_height
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

def set_header_font_size(sheet, size=14):
    for cell in sheet[1]:
        cell.font = Font(size=size, bold=True)

def merge_series_titles(sheet, df):
    col_idx = df.columns.get_loc("Series Title") + 1
    col_letter = get_column_letter(col_idx)
    current_title, start_row = None, 2

    for idx, value in enumerate(df["Series Title"], start=2):
        if value != current_title:
            if current_title is not None and start_row != idx - 1:
                sheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{idx - 1}")
            current_title = value
            start_row = idx
    # Final merge
    if start_row != idx:
        sheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{idx}")

def color_status_cells(sheet, df, columns, positive_value="Yes"):
    for col in columns:
        col_idx = df.columns.get_loc(col) + 1
        for row in range(2, len(df) + 2):
            cell = sheet.cell(row=row, column=col_idx)
            cell.fill = green_fill if cell.value == positive_value else red_fill

def format_date_columns(sheet, df, columns):
    for col in columns:
        col_idx = df.columns.get_loc(col) + 1
        for row in range(2, len(df) + 2):
            sheet.cell(row=row, column=col_idx).number_format = 'DD-MM-YYYY'

# === Main Functions ===

def format_series_sheet(sheet, df, cover_path):
    apply_fill_conditionally(sheet, df, ["Read Count", "Owned Count"])
    add_thumbnails(sheet, df, cover_path)
    set_sheet_dimensions(sheet, df)
    set_header_font_size(sheet)

def format_issues_sheet(sheet, df):
    merge_series_titles(sheet, df)
    color_status_cells(sheet, df, ["Owned Already?", "Read Already?"])
    format_date_columns(sheet, df, ["Bought Date", "Read Date"])
    set_sheet_dimensions(sheet, df, row_height=20)
    set_header_font_size(sheet)

def create_stats_sheet(wb, df_series, df_issues):
    sheet = wb.create_sheet("Stats")

    if df_series.empty and df_issues.empty:
        sheet['A1'] = "No series or issues data available for export."
        return
    
    # === Summary Statistics ===
    total_series = len(df_series)
    total_issues = len(df_issues)
    total_owned = (df_issues["Owned Already?"] == "Yes").sum()
    total_read = (df_issues["Read Already?"] == "Yes").sum()

    percent_owned = round((total_owned / total_issues) * 100, 2) if total_issues else 0
    percent_read = round((total_read / total_issues) * 100, 2) if total_issues else 0

    stats_data = [
        ["Metric", "Value"],
        ["Total Series", total_series],
        ["Total Issues", total_issues],
        ["Owned Issues", total_owned],
        ["Read Issues", total_read],
        ["% Owned", f"{percent_owned}%"],
        ["% Read", f"{percent_read}%"],
    ]

    for r_idx, row in enumerate(stats_data, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, size=12)

    # === Pie Chart: Read vs Unread ===
    pie = PieChart()
    pie.title = "Read vs Unread"
    pie_data = Reference(sheet, min_col=2, min_row=4, max_row=5)  # Read / Unread
    pie_labels = Reference(sheet, min_col=1, min_row=4, max_row=5)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_labels)
    sheet.add_chart(pie, "D2")

    # === Pie Chart: Owned vs Unowned ===
    pie2 = PieChart()
    pie2.title = "Owned vs Unowned"
    owned = total_owned
    unowned = total_issues - total_owned
    sheet["A14"] = "Owned"
    sheet["B14"] = owned
    sheet["A15"] = "Unowned"
    sheet["B15"] = unowned

    pie2_data = Reference(sheet, min_col=2, min_row=14, max_row=15)
    pie2_labels = Reference(sheet, min_col=1, min_row=14, max_row=15)
    pie2.add_data(pie2_data, titles_from_data=False)
    pie2.set_categories(pie2_labels)
    sheet.add_chart(pie2, "D14")

    # === Bar Chart: Top 5 series by owned count ===
    df_series["Owned Number"] = df_series["Owned Count"].apply(parse_owned_count)
    top5 = df_series.sort_values("Owned Number", ascending=False).head(5)

    start_row = 26
    sheet.cell(row=start_row, column=1).value = "Series Title"
    sheet.cell(row=start_row, column=2).value = "Owned"
    sheet.cell(row=start_row, column=1).font = Font(bold=True)
    sheet.cell(row=start_row, column=2).font = Font(bold=True)

    for idx, row in enumerate(top5.itertuples(index=False), start=start_row+1):
        sheet.cell(row=idx, column=1).value = row[1]  # Series Title
        sheet.cell(row=idx, column=2).value = row[-1]  # Owned Number

    bar = BarChart()
    bar.title = "Top 5 Series by Owned Issues"
    bar_data = Reference(sheet, min_col=2, min_row=start_row, max_row=start_row + len(top5))
    bar_labels = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(top5))
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_labels)
    bar.x_axis.title = "Series"
    bar.y_axis.title = "Owned Count"
    bar.width = 16
    bar.height = 8
    sheet.add_chart(bar, "D26")
    set_sheet_dimensions(sheet, df_series, row_height=20)