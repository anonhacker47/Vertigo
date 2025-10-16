import os
from apifairy.decorators import other_responses
from flask import Blueprint, abort, current_app, jsonify,request, send_file
from apifairy import authenticate, body, response

from api import db
from api.models.user import User
from api.models.series import Series
import pandas as pd
from io import BytesIO
from flask import send_file
from api.schemas.user_schema import UserSchema, UpdateUserSchema
from sqlalchemy.orm import selectinload

from api.utils.auth import token_auth
from api.helpers.thumbnail_processing import download_series_thumbnail, save_series_thumbnail
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from api.helpers.excel_processing import create_stats_sheet, format_series_sheet, format_issues_sheet
from babel.numbers import get_currency_symbol

users = Blueprint('users', __name__)
user_schema = UserSchema()
users_schema = UserSchema(many=True)
update_user_schema = UpdateUserSchema(partial=True)


@users.route('/users', methods=['POST'])
@body(user_schema)
@response(user_schema, 201)
def new(args):
    """Register a new user"""
    user = User(**args)
    db.session.add(user)
    db.session.commit()
    return user


# @users.route('/users', methods=['GET'])
# @authenticate(token_auth)
# @paginated_response(users_schema)
# def all():
#     """Retrieve all users"""
#     return User.select()


@users.route('/users/<int:id>', methods=['GET'])
@authenticate(token_auth)
@response(user_schema)
@other_responses({404: 'User not found'})
def get(id):
    """Retrieve a user by id"""
    return db.session.get(User, id) or abort(404)


# @users.route('/users/<username>', methods=['GET'])
# @authenticate(token_auth)
# @response(user_schema)
# @other_responses({404: 'User not found'})
# def get_by_username(username):
#     """Retrieve a user by username"""
#     return db.session.scalar(User.select().filter_by(username=username)) or \
#         abort(404)


@users.route('/me', methods=['GET'])
@authenticate(token_auth)
@response(user_schema)
def me():
    """Retrieve the authenticated user"""
    return token_auth.current_user()


@users.route('/me', methods=['PUT'])
@authenticate(token_auth)
# @body(update_user_schema)
@response(user_schema)
def put():
    """Edit user information (multipart/form-data supported)"""
    user = token_auth.current_user()
    form = request.form

    # Text field updates
    if 'username' in form:
        user.username = form.get('username').strip()
    if 'email' in form:
        user.email = form.get('email').strip()
    if 'preferred_currency' in form:
        currency = form.get('preferred_currency').strip().upper()
        if len(currency) != 3 or not currency.isalpha():
            abort(400, description="Invalid currency code")
        user.preferred_currency = currency

    # Handle password change
    if 'password' in form:
        if 'old_password' not in form or not user.verify_password(form.get('old_password')):
            abort(400, description="Old password missing or incorrect")
        user.password = form.get('password')

    # Handle profile picture
    profile_picture = form.get('profile_picture', '').strip() if 'profile_picture' in form else ''
    if profile_picture.startswith('http'):
        picture_filename = download_series_thumbnail(profile_picture, user.username,'user_path')
        if picture_filename:
            user.profile_picture = picture_filename
    elif 'profile_picture' in request.files:
        file = request.files['profile_picture']
        picture_filename = save_series_thumbnail(file, user.username,'user_path')
        if picture_filename:
            user.profile_picture = picture_filename

    db.session.commit()
    return user

@users.route('me/profile-picture/',methods=['GET'])
@authenticate(token_auth)
def get_profile_picture():
    """Retrieve the User profile picture"""
    user = token_auth.current_user()

    if user is None:
        return jsonify("User not found"), 404

    if user.profile_picture is None:
        return jsonify("noimage")

    try:
        return send_file(os.path.join(current_app.config['user_path'],user.profile_picture))
    except FileNotFoundError:
        return jsonify("Image file not found"), 404
    except Exception as e:
        # Handle other potential exceptions (e.g., permission errors)
        return jsonify(f"Error retrieving image: {str(e)}"), 500
   


@users.route('/export_data', methods=['GET'])
@authenticate(token_auth)
def export_data():
    """Export user data"""
    user = token_auth.current_user()
    
    # Fetch user's series
    series_stmt = user.series_select().options(selectinload(Series.issue))
    series_list = db.session.execute(series_stmt).scalars().all()   
    
    cover_path = current_app.config['cover_path']

    # Prepare data for DataFrame
    series_data = []
    issues_data = []

    currency_code = user.preferred_currency  # e.g., 'USD', 'INR', etc.
    currency_symbol = get_currency_symbol(currency_code)
    purchase_field = f"Purchase Cost ({currency_symbol})"

    for s in series_list:
        missing_issues = [str(i.number) for i in s.issue if not i.is_owned]
        series_data.append({
            "Thumbnail": s.thumbnail,
            "Series Title": s.title,
            "Publisher": ", ".join(p.title for p in s.publisher),
            "Genre": ", ".join(g.title for g in s.genre),
            "Creators": ", ".join(c.title for c in s.creator),
            "Description": s.description,
            "Issue Count": s.issue_count,
            "Owned Count": f"{s.owned_count} / {s.issue_count}",
            "Missing Issues": ", ".join(missing_issues) if missing_issues else "None",
            "Read Count": f"{s.read_count} / {s.issue_count}",
            "Rating": s.user_rating,
            purchase_field: s.purchase_cost,
        })

        for i in s.issue:
            issues_data.append({
                "Series Title": s.title,
                "Issue Number": i.number,
                "Owned Already?": "Yes" if i.is_owned else "No",
                "Read Already?": "Yes" if i.is_read else "No",
                "Bought Price": i.bought_price,
                "Bought Date": i.bought_date,
                "Read Date": i.read_date,
            })


    # Step 2: Create Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_series = pd.DataFrame(series_data)
        df_issues = pd.DataFrame(issues_data)

        # Write dataframes
        df_series.to_excel(writer, index=False, sheet_name='Series')
        df_issues.to_excel(writer, index=False, sheet_name='Issues')

        # Apply formatting only if data exists
        if not df_series.empty:
            format_series_sheet(writer.book['Series'], df_series, cover_path)
        if not df_issues.empty:
            format_issues_sheet(writer.book['Issues'], df_issues)
            
        create_stats_sheet(writer.book, df_series, df_issues)

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"comic_export_user_{user.id}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@users.route('/delete_all_data', methods=['DELETE'])
@authenticate(token_auth)
def delete_all_user_data():
    """Delete all series and issues for the current user"""
    user = token_auth.current_user()

    # Get all series belonging to the user
    series_stmt = user.series_select().options(selectinload(Series.issue))
    series_list = db.session.execute(series_stmt).scalars().all()

    # Delete each series (cascades will handle issues)
    for s in series_list:
        db.session.delete(s)

    db.session.commit()

    return jsonify({"message": f"Deleted {len(series_list)} series and associated issues."}), 200



# @users.route('/me/following', methods=['GET'])
# @authenticate(token_auth)
# @paginated_response(users_schema, order_by=User.username)
# def my_following():
#     """Retrieve the users the logged in user is following"""
#     user = token_auth.current_user()
#     return user.following_select()


# @users.route('/me/followers', methods=['GET'])
# @authenticate(token_auth)
# @paginated_response(users_schema, order_by=User.username)
# def my_followers():
#     """Retrieve the followers of the logged in user"""
#     user = token_auth.current_user()
#     return user.followers_select()


# @users.route('/me/following/<int:id>', methods=['GET'])
# @authenticate(token_auth)
# @response(EmptySchema, status_code=204,
#           description='User is followed.')
# @other_responses({404: 'User is not followed'})
# def is_followed(id):
#     """Check if a user is followed"""
#     user = token_auth.current_user()
#     followed_user = db.session.get(User, id) or abort(404)
#     if not user.is_following(followed_user):
#         abort(404)
#     return {}


# @users.route('/me/following/<int:id>', methods=['POST'])
# @authenticate(token_auth)
# @response(EmptySchema, status_code=204,
#           description='User followed successfully.')
# @other_responses({404: 'User not found', 409: 'User already followed.'})
# def follow(id):
#     """Follow a user"""
#     user = token_auth.current_user()
#     followed_user = db.session.get(User, id) or abort(404)
#     if user.is_following(followed_user):
#         abort(409)
#     user.follow(followed_user)
#     db.session.commit()
#     return {}


# @users.route('/me/following/<int:id>', methods=['DELETE'])
# @authenticate(token_auth)
# @response(EmptySchema, status_code=204,
#           description='User unfollowed successfully.')
# @other_responses({404: 'User not found', 409: 'User is not followed.'})
# def unfollow(id):
#     """Unfollow a user"""
#     user = token_auth.current_user()
#     unfollowed_user = db.session.get(User, id) or abort(404)
#     if not user.is_following(unfollowed_user):
#         abort(409)
#     user.unfollow(unfollowed_user)
#     db.session.commit()
#     return {}


# @users.route('/users/<int:id>/following', methods=['GET'])
# @authenticate(token_auth)
# @paginated_response(users_schema, order_by=User.username)
# @other_responses({404: 'User not found'})
# def following(id):
#     """Retrieve the users this user is following"""
#     user = db.session.get(User, id) or abort(404)
#     return user.following_select()


# @users.route('/users/<int:id>/followers', methods=['GET'])
# @authenticate(token_auth)
# @paginated_response(users_schema, order_by=User.username)
# @other_responses({404: 'User not found'})
# def followers(id):
#     """Retrieve the followers of the user"""
#     user = db.session.get(User, id) or abort(404)
#     return user.followers_select()
